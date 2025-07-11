import openpyxl
from openpyxl import Workbook
import os

# Configuración de archivos
archivo_lista = 'archivo-lista.xlsx'
archivo_plantilla = 'archivo-plantilla.xlsx'
carpeta_salida = 'archivos_generados'

# Crear carpeta de salida si no existe
if not os.path.exists(carpeta_salida):
    os.makedirs(carpeta_salida)

# Cargar archivos
wb_lista = openpyxl.load_workbook(archivo_lista)
wb_plantilla = openpyxl.load_workbook(archivo_plantilla)

# Asumimos que los datos están en la primera hoja
hoja_lista = wb_lista.active
hoja_plantilla = wb_plantilla.active

# Leer datos de la lista
for fila in hoja_lista.iter_rows(min_row=2, values_only=True):  # asumiendo que la primera fila es encabezado
    nombre, apellido, serie = fila[0], fila[1], fila[2]  # ajusta los índices según tu estructura
    
    # Crear una copia de la plantilla en memoria
    nuevo_wb = openpyxl.Workbook()
    nuevo_wb = openpyxl.load_workbook(archivo_plantilla)
    nueva_hoja = nuevo_wb.active
    
    # Aquí debes modificar la plantilla con los datos
    # Ejemplo (ajusta las celdas según tu plantilla real):
    nueva_hoja['A1'] = nombre      # Celda donde va el nombre
    nueva_hoja['B1'] = apellido    # Celda donde va el apellido
    nueva_hoja['C1'] = serie       # Celda donde va la serie
    
    # Guardar el nuevo archivo
    nombre_archivo = f"{carpeta_salida}/{nombre}_{apellido}_{serie}.xlsx"
    nuevo_wb.save(nombre_archivo)
    print(f"Archivo creado: {nombre_archivo}")

print("Proceso completado. Todos los archivos han sido generados.")
