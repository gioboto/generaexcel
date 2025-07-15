import os
from openpyxl import load_workbook
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter

# Carpeta con los archivos Excel generados
carpeta_excel = 'archivos_generados'
carpeta_pdf = 'archivos_pdf'

# Crear carpeta PDF si no existe
os.makedirs(carpeta_pdf, exist_ok=True)

def excel_to_pdf_multiplatform(input_excel_path, output_pdf_path):
    # Leer el archivo Excel con openpyxl
    wb = load_workbook(input_excel_path)
    ws = wb.active  # Obtener la hoja activa
    
    # Crear PDF
    c = canvas.Canvas(output_pdf_path, pagesize=letter)
    width, height = letter
    
    # Configuración inicial
    x_offset = 40
    y_offset = height - 40
    line_height = 14
    
    # Escribir datos de Excel en PDF
    for row in ws.iter_rows(values_only=True):
        line = " | ".join([str(cell) if cell is not None else "" for cell in row])
        
        # Verificar si necesitamos nueva página
        if y_offset <= 40:
            c.showPage()
            y_offset = height - 40
        
        c.drawString(x_offset, y_offset, line)
        y_offset -= line_height
    
    c.save()

# Convertir todos los archivos Excel a PDF
for archivo in os.listdir(carpeta_excel):
    if archivo.endswith('.xlsx'):
        excel_path = os.path.join(carpeta_excel, archivo)
        pdf_name = os.path.splitext(archivo)[0] + '.pdf'
        pdf_path = os.path.join(carpeta_pdf, pdf_name)
        
        excel_to_pdf_multiplatform(excel_path, pdf_path)
        print(f"Convertido: {archivo} → {pdf_name}")

print("Conversión completada!")