import pandas as pd
import os
from openpyxl import load_workbook
from copy import copy
from datetime import datetime 

# PAra la fecha actual
elhoy = datetime.now()
fecha_formateada = elhoy.strftime("%Y/%m/%d")
#fecha_formateada = "2025/07/11"
cadenafechaformateada = str(fecha_formateada)

# Configuración
archivo_lista = 'lista4.xlsx'
archivo_plantilla = 'GTImantenimientoo.xlsx'
carpeta_salida = 'archivos_generados'

# Crear carpeta de salida
os.makedirs(carpeta_salida, exist_ok=True)

# Leer lista de datos
df = pd.read_excel(archivo_lista)

# Cargar plantilla
wb_plantilla = load_workbook(archivo_plantilla)
hoja_plantilla = wb_plantilla.active

for index, row in df.iterrows():
    nombre = row['NOMBRE'] 
    marca = row['MARCA']  
    serie = row['SERIAL']  
    
    # Crear copia de la plantilla
    nuevo_wb = load_workbook(archivo_plantilla)
    nueva_hoja = nuevo_wb.active
    
    # Modificar plantilla 
    nueva_hoja['D5'] = nombre
    nueva_hoja['H5'] = str(marca) + ' / ' + str(serie)
    nueva_hoja['D6'] = cadenafechaformateada
    
    
    # Guardar archivo
    nombre_archivo = f"{carpeta_salida}/{nombre}_{marca}_{serie}.xlsx"
    nuevo_wb.save(nombre_archivo)
    print(f"Archivo generado: {nombre_archivo}")

print("Proceso completado!")
